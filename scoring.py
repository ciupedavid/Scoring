import json

from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment
import openpyxl.utils.cell

with open(r'inAMPNoise.json') as d:
    paths = json.load(d)['Nimble'][0]

project_location = paths['project_location']
device = paths['device']
excel_path = project_location + '\\' + device + '\\' + 'AD8132_WithScores.xlsx'
data_sheet = paths['data_sheet']
gain_sheet = paths['gain_sheet']
score_sheet_name = paths['score_sheet_name']
device_freq = paths['gain_sheet']+' freq'
device_mag = paths['gain_sheet']+' mag'

xl = openpyxl.load_workbook(excel_path)
xl.create_sheet(score_sheet_name)
xl.create_sheet(title='Formatted Data')

sheet1 = xl[data_sheet]
sheet2 = xl[score_sheet_name]
sheet3 = xl[gain_sheet]
sheet4 = xl['Formatted Data']

ColName = {}
Current = 0
for COL in sheet1.iter_cols(1, sheet1.max_column):
    ColName[COL[0].value] = Current
    Current += 1

for freq in sheet1.iter_rows(min_row=2, max_row=100):
    frequency = freq[ColName[device_freq]].value
    sheet2.append({'A': f'{frequency}'})
    # print('%s: cell.value=%s' % (freq[0], freq[0].value))
    magnitude = freq[ColName[device_mag]].value
    sheet4.append({'B': f'{magnitude}'})

move = []

for i in range(1, 100, 1):
    move.append(sheet4.cell(row=i, column=2).value)

for i in range(1, 100, 1):

    for j in range(1, 100, 1):
        sheet2.cell(row=j, column=2).value = move[j - 1]

del xl['Formatted Data']

sheet2.move_range("A1:A100", rows=2, cols=2, translate=True)
sheet2.move_range("B1:B100", rows=2, cols=2, translate=True)

sheet2.delete_rows(3, 1)

last_cell = 100

for col in range(3, sheet2.max_column+1):
    for row in range(1, 100, 1):
        sheet2.cell(column=col, row=row).number_format = '0.00E+00'

xl.active = xl[score_sheet_name]
xl.active.merge_cells('A1:D1')
cell = xl.active.cell(row=1, column=1)
cell.value = 'Info for score'
cell.font = Font(bold=True)
cell.alignment = Alignment(horizontal='center', vertical='center')

xl.active = xl[score_sheet_name]
xl.active.merge_cells('E1:L1')
cell2 = xl.active.cell(row=2, column=1)
xl.active['E1'] = 'Nimble score'
xl.active['E1'].font = Font(bold=True)
cell2.alignment = Alignment(horizontal='center', vertical='center')

xl.active['A2'].alignment = Alignment(wrap_text=True)
xl.active['A2'] = 'Magnitude \n range'
xl.active['B2'].alignment = Alignment(wrap_text=True)
xl.active['B2'] = 'Frequency \n range'
xl.active['C2'].alignment = Alignment(wrap_text=True)
xl.active['C2'] = 'Datasheet \n freq'
xl.active['D2'].alignment = Alignment(wrap_text=True)
xl.active['D2'] = 'Datasheet \n mag'

xl.active['E2'].alignment = Alignment(wrap_text=True)
xl.active['E2'] = 'Closest match \n without going \n over index'
xl.active['F2'].alignment = Alignment(wrap_text=True)
xl.active['F2'] = 'Below \n freq'
xl.active['G2'].alignment = Alignment(wrap_text=True)
xl.active['G2'] = 'Above \n freq'
xl.active['H2'].alignment = Alignment(wrap_text=True)
xl.active['H2'] = 'Below \n mag'
xl.active['I2'].alignment = Alignment(wrap_text=True)
xl.active['I2'] = 'Above \n mag'
xl.active['J2'].alignment = Alignment(wrap_text=True)
xl.active['J2'] = 'Linear \n interpolation'
xl.active['K2'].alignment = Alignment(wrap_text=True)
xl.active['K2'] = 'Error (dB)'
xl.active['L2'] = 'Score'
xl.active['L2'].font = Font(bold=True)

xl.active['A3'] = paths['x_axis_min']
xl.active['A4'] = paths['y_axis_min']

xl.active['B3'] = paths['y_axis_max']
xl.active['B4'] = paths['x_axis_max']

xl.active.column_dimensions['A']. width = 16
xl.active.column_dimensions['B']. width = 16
xl.active.column_dimensions['C']. width = 16
xl.active.column_dimensions['D']. width = 16
xl.active.column_dimensions['E']. width = 16
xl.active.column_dimensions['F']. width = 16
xl.active.column_dimensions['G']. width = 16
xl.active.column_dimensions['H']. width = 16
xl.active.column_dimensions['I']. width = 16
xl.active.column_dimensions['J']. width = 16
xl.active.column_dimensions['K']. width = 16
xl.active.column_dimensions['L']. width = 16

from_cell = sheet3.cell(row=1, column=1)
to_cell = sheet3.cell(row=500, column=1)
cell_range = f'{from_cell.coordinate}:{to_cell.coordinate}'
sheet_name = f'{gain_sheet}!'

for match in range(2, 500):
    match_formula1 = sheet3.cell(row=match, column=1).value
    match_formula2 = sheet2.cell(row=match+1, column=3).value
    final_match_formula = f'=MATCH({match_formula2},{sheet_name}{cell_range},1)'
    sheet2.cell(row=match+1, column=5).value = final_match_formula

sheet_name_index = f'{gain_sheet}!'

index_from_cell = sheet2.cell(row=3, column=5)
index_to_cell = sheet3.cell(row=100, column=5)
index_cell_range = f'{index_from_cell.coordinate}:{index_to_cell.coordinate}'

for index in range(2, 100):
    final_index_formula = f'=INDEX({sheet_name_index}{cell_range},{index_cell_range})'
    sheet2.cell(row=index+1, column=6).value = final_index_formula

for index2 in range(2, 100):
    final_index_formula2 = f'=INDEX({sheet_name_index}{cell_range},{index_cell_range}+1)'
    sheet2.cell(row=index2+1, column=7).value = final_index_formula2

from_cell = sheet3.cell(row=1, column=2)
to_cell = sheet3.cell(row=500, column=2)
cell_range2 = f'{from_cell.coordinate}:{to_cell.coordinate}'

for index3 in range(2, 100):
    final_index_formula3 = f'=INDEX({sheet_name_index}{cell_range2},{index_cell_range})'
    sheet2.cell(row=index3+1, column=8).value = final_index_formula3

for index4 in range(2, 100):
    final_index_formula4 = f'=INDEX({sheet_name_index}{cell_range2},{index_cell_range}+1)'
    sheet2.cell(row=index4+1, column=9).value = final_index_formula4

for slope in range(2, 100):
    datasheet_freq = sheet2.cell(row=slope + 1, column=3)
    below_freq = sheet2.cell(row=slope + 1, column=6)
    above_freq = sheet2.cell(row=slope + 1, column=7)
    below_mag = sheet2.cell(row=slope + 1, column=8)
    above_mag = sheet2.cell(row=slope + 1, column=9)
    final_slope = f'=SLOPE({below_mag.coordinate}:{above_mag.coordinate},{below_freq.coordinate}:{above_freq.coordinate})*({datasheet_freq.coordinate}-{below_freq.coordinate})+{below_mag.coordinate}'
    sheet2.cell(row=slope+1, column=10).value = final_slope

for abs in range(2, 100):
    linear_interpolation = sheet2.cell(row=abs + 1, column=10)
    data_sheet = sheet2.cell(row=abs + 1, column=4)
    final_abs = f'=ABS({linear_interpolation.coordinate}-{data_sheet.coordinate})'
    sheet2.cell(row=abs+1, column=11).value = final_abs

for average in range(2, 100):
    error_min = sheet2.cell(row=3, column=11)
    error_max = sheet2.cell(row=100, column=11)
    error = f'{error_min.coordinate}:{error_max.coordinate}'
    final_error = f'=AVERAGE({error})'
    sheet2.cell(row=3, column=12).value = final_error
    xl.active['L3'].font = Font(bold=True)
    xl.active['E1'].alignment = Alignment(horizontal='center', vertical='center')

for rows in sheet2.iter_rows(min_row=1, max_row=100, min_col=1, max_col=4):
    for cell in rows:
        cell.fill = PatternFill(start_color='FFF2CC', fill_type='solid')

for rows2 in sheet2.iter_rows(min_row=1, max_row=100, min_col=5, max_col=12):
    for cell2 in rows2:
        cell2.fill = PatternFill(start_color='E2EFDA', fill_type='solid')

sheet2.delete_rows(60, 500)
xl.save(excel_path)